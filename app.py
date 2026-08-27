import streamlit as st
import pandas as pd
import gspread
from google.oauth2 import service_account
from io import BytesIO
import openpyxl

st.set_page_config(page_title="X5 Торги — Автозаполнение", layout="centered")
st.title("📦 Автозаполнение ставок для торгов X5")

# === 1. Загрузка справочника из Google Sheets (новая структура) ===
@st.cache_data(ttl=3600)
def load_prices_from_gsheets():
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = service_account.Credentials.from_service_account_info(
        st.secrets["gcp_service_account"], scopes=scope
    )
    client = gspread.authorize(creds)
    sheet = client.open_by_key("17q72JGZmLfGcKjMcWhj8RBdnq06dsUFkmeZi6kuaXpA").worksheet("Цены")
    
    all_values = sheet.get_all_values()
    # Данные начинаются с 4-й строки (индекс 3)
    data_rows = all_values[3:]
    
    # Выбираем нужные колонки:
    # B (индекс 1) – PLU
    # D (индекс 3) – Страна
    # I, K, M, O, Q, S, U, W, Y, AA, AC, AE, AG, AI
    # индексы: 8,10,12,14,16,18,20,22,24,26,28,30,32,34
    price_indices = [8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30, 32, 34]
    price_columns = ["I", "K", "M", "O", "Q", "S", "U", "W", "Y", "AA", "AC", "AE", "AG", "AI"]
    
    selected = []
    for row in data_rows:
        # Если строка короче 35, дополняем пустыми (макс индекс 34)
        if len(row) < 35:
            row = row + [''] * (35 - len(row))
        # Собираем PLU, Страну и цены
        row_data = [row[1], row[3]]  # PLU, Страна
        for idx in price_indices:
            row_data.append(row[idx] if idx < len(row) else '')
        selected.append(row_data)
    
    col_names = ["PLU", "Страна"] + price_columns
    df = pd.DataFrame(selected, columns=col_names)
    df = df[df["PLU"].notna() & (df["PLU"] != "")]
    df["PLU"] = df["PLU"].astype(str).str.strip()
    
    # Преобразуем ценовые колонки в числа
    for col in price_columns:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    
    return df

# === 2. Справочник РЦ → колонка (обновлённый) ===
RC_TO_COLUMN = {
    "РЦ Софьино ФРОВ": "Q",
    "РЦ Х Воронеж": "S",
    "РЦ Рамонь-Алкоголь": "I",
    "РЦ 5 Курск-Алкоголь": "I",
    "РЦ 5 Тамбов": "I",
    "РЦ СЛК": "W",
    "РЦ 5 Самара": "K",
    "РЦ Кузнецк-Алкоголь": "K",
    "РЦ Саратов-Алкоголь": "K",
    "РЦ 5 Волгоград": "K",
    "Группа РЦ: РЦ Кузнецк-Алкоголь,РЦ Саратов-Алкоголь": "K",
    "РЦ Санкт-Петербург": "U",
    "РЦ Х Нижний Новгород": "S",
    "РЦ 5 Южный Адыгея": "M",
    "РЦ 5 Невинномысск-Алкоголь": "M",
    "РЦ 5 Краснодар": "M",
    "Группа РЦ: РЦ 5 Южный Адыгея,РЦ 5 Краснодар": "M",
    "РЦ 5 Ростов Алкоголь 2": "M",
    "РЦ Х Адыгея": "W",
    "РЦ УФА Сигма-Алкоголь": "O",
    "РЦ 5 Оренбург Север": "O",
    "РЦ 5 Пермь 2": "O",
    "РЦ Падиково": "Y",
    "РЦ Литвиново": "Y",
    "РЦ Валищево": "Y",
    "РЦ Купавна": "Y",
    "РЦ Воронеж": "Q",
    "2PL РЦ Воронеж Холодный": "AA",
    "РЦ Дзержинск": "AA",
    "РЦ Ярославль": "AA",
    "Группа РЦ: РЦ Воронеж,2PL РЦ Воронеж Холодный": "AA",
    "РЦ Казань": "AC",
    "РЦ Самара": "AC",
    "РЦ Саратов": "AC",
    "РЦ Ростов-на-Дону": "AE",
    "РЦ Краснодар": "AE",
    "РЦ Волгоград": "AE",
    "РЦ Пермь": "AG",
    "РЦ Уфа": "AG",
    "РЦ Екатеринбург": "AG",
    "РЦ Челябинск": "AG",
    "РЦ Кемерово": "AI",
    "РЦ Новосибирск Садовый": "AI",
}

def get_price_column_for_rc(rc_name):
    rc_name = rc_name.strip()
    for key, col in RC_TO_COLUMN.items():
        if key.lower() == rc_name.lower():
            return col
    return None

# === 3. Загрузка файла ===
uploaded_file = st.file_uploader(
    "Загрузите файл с торгами (Excel)",
    type=["xlsx", "xls"],
    help="Скачайте файл с портала Х5 и загрузите его сюда"
)

# === 4. Параметры ===
col1, col2 = st.columns(2)
with col1:
    comment = st.text_input("📝 Комментарий (общий для всех позиций)", placeholder="Оставьте пустым, чтобы не менять")
with col2:
    discount = st.number_input("💰 Скидка (руб.)", min_value=0.0, step=0.5, value=0.0, help="0 — без скидки")

# === 5. Обработка ===
if uploaded_file is not None and st.button("🚀 Обработать файл"):
    with st.spinner("Идёт обработка..."):
        try:
            prices_df = load_prices_from_gsheets()
        except Exception as e:
            st.error(f"Ошибка при загрузке Google-таблицы: {e}")
            st.stop()

        try:
            df = pd.read_excel(uploaded_file, sheet_name=0)
        except Exception as e:
            st.error(f"Ошибка при чтении файла: {e}")
            st.stop()

        # Приводим колонки к нужным типам
        df["Страна"] = df["Страна"].astype(str)
        df["Мой комментарий"] = df["Мой комментарий"].astype(str)
        df["Мое предложение"] = pd.to_numeric(df["Мое предложение"], errors="coerce")
        df["Мой гарантированный объем"] = pd.to_numeric(df["Мой гарантированный объем"], errors="coerce")

        # Временные колонки
        df["Цена_из_справочника"] = None
        df["Страна_из_справочника"] = None
        unknown_rcs = set()

        for idx, row in df.iterrows():
            plu = str(row["Код PLU"]).strip()
            rc = row["РЦ доставки"]

            price_row = prices_df[prices_df["PLU"] == plu]
            if price_row.empty:
                continue

            col_letter = get_price_column_for_rc(rc)
            if col_letter is None:
                unknown_rcs.add(rc)
                continue

            price_value = price_row.iloc[0][col_letter]
            if pd.notna(price_value):
                df.at[idx, "Цена_из_справочника"] = float(price_value)
                country = price_row.iloc[0]["Страна"]
                if pd.isna(country) or country is None:
                    country = ""
                df.at[idx, "Страна_из_справочника"] = str(country)

        if unknown_rcs:
            st.warning(f"⚠️ Для следующих РЦ не найдено соответствие: {', '.join(unknown_rcs)}")

        updated_count = 0
        for idx, row in df.iterrows():
            if pd.notna(row.get("Цена_из_справочника")):
                new_price = row["Цена_из_справочника"] - discount
                df.at[idx, "Мое предложение"] = round(new_price, 2)
                # Столбец M (самовывоз) не трогаем
                df.at[idx, "Страна"] = row["Страна_из_справочника"]
                if comment.strip():
                    df.at[idx, "Мой комментарий"] = comment.strip()
                df.at[idx, "Мой гарантированный объем"] = row["Количество"]
                updated_count += 1

        df = df.drop(columns=["Цена_из_справочника", "Страна_из_справочника"], errors="ignore")

        # Удаляем строки, где нет цены или страны
        before = len(df)
        df = df[df["Мое предложение"].notna() & df["Страна"].notna()]
        after = len(df)
        deleted = before - after

        if df.empty:
            st.warning("⚠️ После удаления пустых позиций не осталось ни одной строки.")
            st.stop()

        # Заменяем NaN на пустые строки
        df = df.fillna("")

        # Сохраняем в Excel, затем очищаем столбцы F и M через openpyxl
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Сбор предложений")
            workbook = writer.book
            sheet = workbook["Сбор предложений"]

            col_f = None
            col_m = None
            for col_idx, col_name in enumerate(df.columns, start=1):
                if col_name == "Предложение X5 с доставкой":
                    col_f = col_idx
                if col_name == "Мое предложение (самовывоз)":
                    col_m = col_idx

            if col_f:
                for row in range(2, sheet.max_row + 1):
                    cell = sheet.cell(row=row, column=col_f)
                    cell.value = None
            if col_m:
                for row in range(2, sheet.max_row + 1):
                    cell = sheet.cell(row=row, column=col_m)
                    cell.value = None

        output.seek(0)

        st.success(f"✅ Обновлено {updated_count} позиций. Удалено пустых: {deleted}.")
        if deleted > 0:
            st.info(f"ℹ️ Удалены строки, которые не были заполнены ни автоматически, ни вручную.")

        st.download_button(
            label="📥 Скачать файл",
            data=output,
            file_name=uploaded_file.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        with st.expander("👁️ Превью результата (первые 5 строк)"):
            st.dataframe(df.head())